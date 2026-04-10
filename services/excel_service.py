"""
Excel 生成服务
生成店铺分析报告 Excel 文件
"""
import io
import os
from datetime import datetime
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import logging

from services.search_service import SearchResult

logger = logging.getLogger(__name__)


class ExcelService:
    """Excel 生成服务"""

    def __init__(self):
        self.header_font = Font(bold=True, size=12)
        self.title_font = Font(bold=True, size=14)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        self.header_font_white = Font(bold=True, size=11, color='FFFFFF')

    def generate_report(self, search_results: List[SearchResult], analysis_result: str, matched_companies: Dict[str, str] = None) -> io.BytesIO:
        """
        生成分析报告 Excel

        Args:
            search_results: 搜索结果列表
            analysis_result: AI 分析结果
            matched_companies: 匹配的公司 {店铺名: 匹配的公司名}

        Returns:
            Excel 文件流
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "店铺分析报告"

        # 设置列宽 - 增加电话、邮箱列
        column_widths = [18, 30, 10, 10, 18, 12, 30, 15, 25, 15, 30]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # 标题行
        ws.merge_cells('A1:K1')
        ws['A1'] = f"店铺分析报告 - {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws['A1'].font = self.title_font
        ws['A1'].alignment = Alignment(horizontal='center')

        # 表头
        headers = ['店铺名称', '公司名称', '法人', '状态', '注册资本', '成立日期', '地址', '电话', '邮箱', '统一信用代码', '匹配结果']
        row = 3
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font_white
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center')

        # 数据行
        row = 4
        current_shop = None
        shop_start_row = row

        for result in search_results:
            # 如果是新店铺，添加店铺名称行
            if result.shop_name != current_shop:
                if current_shop is not None and row > shop_start_row:
                    # 合并上一个店铺的店铺名称单元格
                    ws.merge_cells(f'A{shop_start_row}:A{row-1}')
                    ws[f'A{shop_start_row}'].alignment = Alignment(vertical='center', horizontal='center')
                current_shop = result.shop_name
                shop_start_row = row

            # 添加公司数据
            if result.company_list:
                for company in result.company_list:
                    ws.cell(row=row, column=1, value=result.shop_name).border = self.border
                    
                    # 公司名称
                    ws.cell(row=row, column=2, value=company.get('name', '')).border = self.border
                    
                    # 法人
                    ws.cell(row=row, column=3, value=company.get('legalPerson', '')).border = self.border
                    
                    # 状态
                    ws.cell(row=row, column=4, value=company.get('regStatus', '')).border = self.border
                    
                    # 注册资本
                    ws.cell(row=row, column=5, value=company.get('capital', '')).border = self.border
                    
                    # 成立日期
                    ws.cell(row=row, column=6, value=company.get('establishDate', '')).border = self.border
                    
                    # 地址
                    ws.cell(row=row, column=7, value=company.get('address', '')).border = self.border
                    
                    # 电话
                    ws.cell(row=row, column=8, value=company.get('phone', '')).border = self.border
                    
                    # 邮箱
                    ws.cell(row=row, column=9, value=company.get('email', '')).border = self.border
                    
                    # 统一信用代码
                    ws.cell(row=row, column=10, value=company.get('creditCode', '')).border = self.border
                    
                    # 匹配结果
                    matched = ''
                    if matched_companies:
                        matched = matched_companies.get(result.shop_name, '')
                        if matched == company.get('name', ''):
                            matched_cell = ws.cell(row=row, column=11, value='✓ 匹配')
                            matched_cell.font = Font(color='00B050', bold=True)
                            matched_cell.border = self.border
                        else:
                            ws.cell(row=row, column=11, value='').border = self.border
                    else:
                        ws.cell(row=row, column=11, value='').border = self.border
                    
                    row += 1
            else:
                # 没有搜索到公司
                ws.cell(row=row, column=1, value=result.shop_name).border = self.border
                ws.cell(row=row, column=2, value='未找到相关企业').border = self.border
                for col in range(3, 12):
                    ws.cell(row=row, column=col, value='').border = self.border
                row += 1

        # 合并最后一个店铺的店铺名称单元格
        if row > shop_start_row:
            ws.merge_cells(f'A{shop_start_row}:A{row-1}')
            ws[f'A{shop_start_row}'].alignment = Alignment(vertical='center', horizontal='center')

        # AI 分析结果（放在最后）
        row += 2
        ws.merge_cells(f'A{row}:K{row}')
        ws.cell(row=row, column=1, value='AI 分析结果').font = self.title_font
        
        row += 1
        ws.merge_cells(f'A{row}:K{row+10}')
        analysis_cell = ws.cell(row=row, column=1, value=analysis_result)
        analysis_cell.alignment = Alignment(wrap_text=True, vertical='top')

        # 保存到内存流
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output


# 全局实例
excel_service = ExcelService()